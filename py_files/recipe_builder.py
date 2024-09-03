#file name : recipe_builder.py

import kivy
from kivy.app import App
from kivy.uix.label import Label
from kivy.lang import Builder
from kivy.uix.textinput import TextInput
from kivy.uix.widget import Widget
from kivymd.uix.boxlayout import BoxLayout
from kivy.properties import ObjectProperty
from kivy.uix.recycleview import RecycleView
from kivymd.app import MDApp
from kivymd.uix.list import IRightBodyTouch,OneLineAvatarIconListItem
from kivymd.uix.selectioncontrol import MDCheckbox
from kivymd.icon_definitions import md_icons
from kivy.properties import StringProperty
from kivymd.uix.list import OneLineListItem
from kivymd.uix.menu import MDDropdownMenu
from kivy.metrics import dp
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDFlatButton
import openpyxl

#kivy.require('2.0.0')


class ListItemWithCheckbox(OneLineAvatarIconListItem):
    pass

class RightCheckbox(IRightBodyTouch, MDCheckbox):
    pass

class Content(BoxLayout):
    pass
class Content1(BoxLayout):
    pass
path = "C:/Users/sloan/OneDrive/Desktop/Python Projects/grocery_items.xlsx"
ingr_list = []
cost_list = []
recipe_save = []
measurements = {"1 cup":1,"1/2 cup":.5,"1/4 cup":.25,"1/3 cup":.33, "1 teaspoon":.167,"1 tablespoon":.05,"1":1,"2":2,"3":3,"4":4,"5":5,"6":6,"7":7}
class ListApp(MDApp):
    dialog = None
    def meas_dropdown(self):     
        self.menu_items = [
            {
                "viewclass": "OneLineListItem",
                "text": f"{d}",
                "on_press":lambda x = f"{d}" : self.set_item(x),
            }for d in measurements
        ]
        #print()
        self.menu = MDDropdownMenu(
            caller = self.root.ids.num_input,
            items = self.menu_items,
            position = "bottom",
            width_mult = 4,
        )
        #print("TESTING")
        self.menu.open()
        

    def set_item(self,text__item):
        #print("PRESSED")
        self.root.ids.num_input.text = text__item
        self.menu.dismiss()

    def ing_recipe_dropdown(self):
        
        path = "C:/Users/sloan/OneDrive/Desktop/Python Projects/grocery_items.xlsx"

        wb_obj = openpyxl.load_workbook(path)

        sheet_obj = wb_obj["Ingredients"]

        m_row = sheet_obj.max_row
        total = 0
        for i in range(1, m_row+1 ):
            cell_obj = sheet_obj.cell(row = i, column = 1)
            cell_obj2 = sheet_obj.cell(row=i, column = 2 )
            #print(cell_obj.value)
            ingr_list.append(cell_obj.value)
            print(cell_obj, cell_obj2)
            #for stuff in ingr_list:
                #print(stuff)
            #total = total + cell_obj2.value
            #print(total)
            #self.root.ids.list_total.text = str(total)
        #for d in ingr_list:
            #print(d)
        self.ing_items = [
            {
                "viewclass": "OneLineListItem",
                "text": f"{i}",
                "on_release":lambda x = f"{i}" : self.set_ing(x),
            }for i in ingr_list
        ]
        #print()
        self.menu = MDDropdownMenu(
            caller = self.root.ids.ing_input,
            items = self.ing_items,
            position = "bottom",
            width_mult = 4,
        )
        self.menu.open()
        
    def set_ing(self,text_item):
        self.root.ids.ing_input.text = text_item
        self.menu.dismiss()


    def build(self):
        self.theme_cls.theme_style = "Light"
        self.theme_cls.primary_palette = "Indigo"
        return Builder.load_file("recipe_builder.kv")

    def recipe_name(self):
        recipename = self.root.ids.entry_input.text
        self.root.ids.recipe_name_label.text = recipename
        self.root.ids.entry_input.text = ''
        return recipename
    
    def running_total(self):
        for i in cost_list:
            run_total =sum(cost_list)
            run_total = round(run_total,2)
            run_total = str(run_total)
        print(run_total)
        self.root.ids.list_total.text.run_total
            

    def ing_name(self):
        flag = 0
        cost = 0
        running_total = 0
        ingredient_name = self.root.ids.ing_input.text
        
        ingredient_number = self.root.ids.num_input.text

        #self.root.ids.ing_input.text = ''
        #self.root.ids.num_input.text = ''

        print(ingredient_number,ingredient_name)

        #if ingredient_name and ingredient_number:

            
        path = "C:/Users/sloan/OneDrive/Desktop/Python Projects/grocery_items.xlsx"

        wb_obj = openpyxl.load_workbook(path)

        sheet_obj = wb_obj.active

        m_row = sheet_obj.max_row
        
        for i in range(1,m_row+1):
            cell_obj_1 = sheet_obj.cell(row = i, column = 1)
            cell_obj_2 = sheet_obj.cell(row=i,column = 2)

            if ingredient_name == cell_obj_1.value:
                for key,value in measurements.items():
                    if key == ingredient_number:
                        conversion = float(value)
                cost = float(cell_obj_2.value)*float(conversion)
                print(cost)
                cost = round(cost,2)
                cost_list.append(cost)
                sing_item = [ingredient_name, int(conversion), cost]
                recipe_save.append(sing_item)
                #list_dict["ingredients"] += [ingredient_name]
                #list_dict["number"] += [ingredient_number]
                #list_dict["cost"] += [cost]\
                
                #print(total)
                self.root.ids.list_container.add_widget(ListItemWithCheckbox(text=f"{ingredient_number}  {ingredient_name}  {cost}"))
                flag = 1
                break
            else:
                flag =0
        #self.root.ids.list_total.text = str(total)
        print(cost_list)
        sum_ = sum(cost_list)
        self.root.ids.list_total.text = str(sum_)

        if flag == 1:
            print("****FOUND")
            
        elif flag == 0:
            #print("****NOT FOUND")
            self.open_dialog(ingredient_name)
                        #self.dialog = None


    def open_dialog(self, ingredient_name):
        #if not self.dialog:
        print("TESTING")
        self.dialog = MDDialog(
            title = "Add An Ingredient",
            type = 'custom',
            #text = f"{ingredient_name} Doesn't Exist - Please Add Item",
            content_cls = Content(),
            buttons = [
                MDFlatButton(
                    
                    text="Add Item",
                    theme_text_color = "Custom",
                    text_color = self.theme_cls.primary_color,
                    on_release = self.add_item
                ),
                MDFlatButton(
                    
                    text = "Cancel",
                    theme_text_color="Custom",
                    text_color = self.theme_cls.primary_color,
                    on_release = self.close_dialog
                ),
            ],
        )
        self.dialog.open()
    
    def testing_item(self):
        print("TESTING")


    def close_dialog(self,obj):
        
        self.dialog.dismiss()

    def add_item(self,obj):
        #print('heare i wan to print the Todo and time')
        new_ingr = self.dialog.content_cls.ids.new_ingredient.text
        new_price = float(self.dialog.content_cls.ids.new_price.text)
        new_weight =  int(self.dialog.content_cls.ids.new_weight.text)
        #new_entry = new_ingr + 
        path = "C:/Users/sloan/OneDrive/Desktop/Python Projects/grocery_items.xlsx"

        wb_obj = openpyxl.load_workbook(path)

        sheet = wb_obj.active
        sheet.append([new_ingr,new_price,new_weight])
        wb_obj.save(path)
        
        #print("TEST")
        self.dialog.dismiss()

                    
    def but_press(self,widget):
        #print(self.root.ids.list_container)
        #print(self.root.ids.list_container.children)
        self.root.ids.list_container.remove_widget(widget)
        #print('delete')
      

    def save_list(self,obj):
        print("TESTING")
        self.dialog.dismiss()
        path = "C:/Users/sloan/OneDrive/Desktop/Python Projects/grocery_items.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        #sheet_obj = wb_obj.active
        final_name = self.root.ids.recipe_name_label.text
        print(final_name)
        
        if f"{final_name} -RECIPE" in wb_obj:
            self.recipe_name_dialog()
            
        else:
            new_list_sheet = wb_obj.create_sheet()
            new_list_sheet.title = f"{final_name} -RECIPE"

        
            for ind in recipe_save:
                new_list_sheet.append(ind)
                print(ind)
                
        
            wb_obj.save(path)
        

    def save_dialog(self):
        #if not self.dialog:
        #print("TESTING")
        self.dialog = MDDialog(
            title = "Are you sure you want to save?",
            text = "Saving will exit Recipe Builder",
            type = 'custom',
            
            content_cls = Content1(),
            buttons = [
                MDFlatButton(
                    
                    text="Finish Recipe",
                    theme_text_color = "Custom",
                    text_color = self.theme_cls.primary_color,
                    on_release = self.save_list,
                ),
                MDFlatButton(
                    
                    text = "Cancel",
                    theme_text_color="Custom",
                    text_color = self.theme_cls.primary_color,
                    on_release = self.close_dialog
                ),
            ],
        )
        self.dialog.open()      

    def recipe_name_dialog(self):
        #if not self.dialog:
        #print("TESTING")
        self.dialog = MDDialog(
            title = "That name already exists - Please choose another",
            type = 'custom',
            
            content_cls = Content(),
            buttons = [
                MDFlatButton(
                    
                    text = "Okay",
                    theme_text_color="Custom",
                    text_color = self.theme_cls.primary_color,
                    on_release = self.close_dialog
                ),
            ],
        )
        self.dialog.open()    
                



if __name__ == '__main__':
    ListApp().run()
