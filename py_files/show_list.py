import kivy
from kivy.app import App
from kivy.uix.label import Label
from kivy.lang import Builder
from kivy.uix.textinput import TextInput
from kivy.uix.widget import Widget
from kivymd.uix.boxlayout import BoxLayout
from kivy.properties import ObjectProperty
from kivy.uix.recycleview import RecycleView
from kivymd.app import MDApp
from kivymd.uix.list import IRightBodyTouch,OneLineAvatarIconListItem
from kivymd.uix.selectioncontrol import MDCheckbox
from kivymd.icon_definitions import md_icons
from kivy.properties import StringProperty
from kivymd.uix.list import OneLineListItem
from kivymd.uix.menu import MDDropdownMenu
from kivy.metrics import dp
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDFlatButton
from kivy.uix.screenmanager import ScreenManager, Screen,WipeTransition
import openpyxl

wb_name = "Test"

class ListItemWithCheckbox(OneLineAvatarIconListItem):
    pass

class RightCheckbox(IRightBodyTouch, MDCheckbox):
    pass
class Content(BoxLayout):
    pass
class ShowListApp(MDApp):
    def build(self):
        self.theme_cls.theme_style = "Light"
        self.theme_cls.primary_palette = "Indigo"
        return Builder.load_file("show_list.kv")
    
    def on_start(self):
        #print('TEST')
        path = "C:/Users/sloan/OneDrive/Desktop/Python Projects/grocery_items.xlsx"

        wb_obj = openpyxl.load_workbook(path)
        #self.root.ids.scroll_list.clear_widgets(widget)
        #print(wb_obj.sheetnames)
        for wb in wb_obj.sheetnames:
            
            if "-LIST" in wb or "-RECIPE" in wb:
                print(wb)
                wb_name = wb
                self.root.ids.show_list_container.add_widget(ListItemWithCheckbox(text=f"{wb}"))
                
    def recipe_popup(self):
        #print(wb_name)
        self.dialog = MDDialog(
            title = "Recipe/List",
            type = 'custom',
            #text = f"{ingredient_name} Doesn't Exist - Please Add Item",
            content_cls = Content(),
            buttons = [
                MDFlatButton(
                    
                    text="Edit",
                    theme_text_color = "Custom",
                    text_color = self.theme_cls.primary_color,
                    on_release = self.testing_item
                ),
                MDFlatButton(
                    
                    text = "Cancel",
                    theme_text_color="Custom",
                    text_color = self.theme_cls.primary_color,
                    on_release = self.close_dialog
                ),
            ],
        )
        self.dialog.open()
    
    def testing_item(self):
        print("TESTING")


    def close_dialog(self,obj):
        
        self.dialog.dismiss()



if __name__ == "__main__":
    ShowListApp().run()

    